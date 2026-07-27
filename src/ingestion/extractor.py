import json
import csv
import pdfplumber
from docx import Document
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from pathlib import Path


def extract_pdf(path: Path) -> str:
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            content = page.extract_text()
            if content:
                pages.append(content)
    return "\n".join(pages)


def extract_docx(path: Path) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_xlsx(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        headers = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            cells = [str(c) if c is not None else "" for c in row]
            if i == 0:
                headers = cells
            elif any(cells):
                lines.append(", ".join(f"{h}: {v}" for h, v in zip(headers, cells) if v))
    return "\n".join(lines)


def extract_html(path: Path) -> str:
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    return soup.get_text(separator="\n", strip=True)


def extract_markdown(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def extract_csv(path: Path) -> str:
    lines = []
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            lines.append(", ".join(f"{k}: {v}" for k, v in row.items() if v))
    return "\n".join(lines)


def extract_json(path: Path) -> str:
    data = json.loads(path.read_text(encoding="utf-8"))
    return json.dumps(data, ensure_ascii=False, indent=2)


EXTRACTORS = {
    ".pdf":  extract_pdf,
    ".docx": extract_docx,
    ".xlsx": extract_xlsx,
    ".html": extract_html,
    ".md":   extract_markdown,
    ".csv":  extract_csv,
    ".json": extract_json,
}

SUPPORTED_EXTENSIONS = set(EXTRACTORS.keys())


def extract(path: Path) -> str | None:
    fn = EXTRACTORS.get(path.suffix.lower())
    if fn is None:
        return None
    return fn(path)
